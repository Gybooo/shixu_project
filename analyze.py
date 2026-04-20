import os
import sys
import csv
import io
import zipfile
from datetime import datetime, timezone

import numpy as np
import pandas as pd
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.dates as mdates

os.environ["HF_HOME"] = r"E:\timesfm_project\hf_cache"

import torch
import timesfm

plt.rcParams['font.sans-serif'] = ['SimHei', 'Microsoft YaHei']
plt.rcParams['axes.unicode_minus'] = False

OUTPUT_DIR = r"E:\timesfm_project\output"
os.makedirs(OUTPUT_DIR, exist_ok=True)

DATA_DIR = r"C:\Users\Administrator\Desktop\时序"
LOG_FILE = os.path.join(DATA_DIR, "log8.12-8.18(1).xlsx")
ZIP_FILE = os.path.join(DATA_DIR, "2024.07.17-08.17.zip")

# ============================================================
# Part 1: Load model
# ============================================================
print("=" * 60)
print("Part 1: Loading TimesFM 2.5 model...")
print("=" * 60)

torch.set_float32_matmul_precision("high")
model = timesfm.TimesFM_2p5_200M_torch.from_pretrained(
    "google/timesfm-2.5-200m-pytorch"
)
model.compile(
    timesfm.ForecastConfig(
        max_context=1024,
        max_horizon=256,
        normalize_inputs=True,
        use_continuous_quantile_head=True,
        force_flip_invariance=True,
        infer_is_positive=True,
        fix_quantile_crossing=True,
    )
)
print("Model loaded successfully!\n")

# ============================================================
# Part 2: Load downtime log
# ============================================================
print("=" * 60)
print("Part 2: Loading downtime log...")
print("=" * 60)

import openpyxl
wb = openpyxl.load_workbook(LOG_FILE, data_only=True)
ws = wb['Sheet1']
log_rows = []
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
    log_rows.append({
        'EqpNo': row[0], 'EqpType': row[1], 'MDTType': row[3],
        'MDTFunction': row[4], 'LogTime': row[6], 'LogState': row[7],
        'LogText': row[8], 'Duration': row[9] if row[9] else 0,
        'UnplanCode': row[10], 'Date': str(row[11])
    })
log_df = pd.DataFrame(log_rows)
log_df['LogTime'] = pd.to_datetime(log_df['LogTime'])
print(f"  Loaded {len(log_df)} log entries")

hourly_downtime = log_df[log_df['LogState'] == 'Unplanned'].copy()
hourly_downtime['hour'] = hourly_downtime['LogTime'].dt.floor('h')
hourly_stats = hourly_downtime.groupby('hour').agg(
    count=('Duration', 'count'),
    total_duration=('Duration', 'sum')
).reset_index()

all_hours = pd.date_range(
    start=log_df['LogTime'].min().floor('h'),
    end=log_df['LogTime'].max().ceil('h'),
    freq='h'
)
hourly_full = pd.DataFrame({'hour': all_hours})
hourly_full = hourly_full.merge(hourly_stats, on='hour', how='left').fillna(0)
print(f"  Hourly downtime stats: {len(hourly_full)} hours\n")

# ============================================================
# Part 3: Load vibration data from ZIP (Aug 12-17, MPB02)
# ============================================================
print("=" * 60)
print("Part 3: Loading vibration data from ZIP (MPB02, Aug 12-14)...")
print("=" * 60)

z = zipfile.ZipFile(ZIP_FILE, 'r')
target_files = sorted([
    n for n in z.namelist()
    if 'MPB02' in n and ('08.12' in n or '08.13' in n or '08.14' in n)
])
print(f"  Found {len(target_files)} files to load")

vibration_data = []
for fname in target_files:
    print(f"    Reading {fname.split('/')[-1]}...")
    with z.open(fname) as zf:
        content = io.TextIOWrapper(zf, encoding='utf-8-sig')
        reader = csv.reader(content)
        for _ in range(4):
            next(reader)
        for row in reader:
            if len(row) >= 9 and row[8] == 'MPB_02' and row[7] == 'Magnitude':
                try:
                    vibration_data.append({
                        'time': row[5],
                        'value': float(row[6])
                    })
                except ValueError:
                    pass

vib_df = pd.DataFrame(vibration_data)
vib_df['time'] = pd.to_datetime(vib_df['time'], format='ISO8601', utc=True).dt.tz_localize(None)
vib_df = vib_df.sort_values('time').reset_index(drop=True)
print(f"  Total vibration records: {len(vib_df)}")
print(f"  Time range: {vib_df['time'].min()} ~ {vib_df['time'].max()}\n")

# Resample to 1-minute mean for manageability
vib_1min = vib_df.set_index('time').resample('1min').mean().dropna()
print(f"  After 1-min resample: {len(vib_1min)} points\n")

# ============================================================
# Part 4: TimesFM vibration forecast
# ============================================================
print("=" * 60)
print("Part 4: Vibration forecast with TimesFM...")
print("=" * 60)

vib_values = vib_1min['value'].values
context_len = min(1024, len(vib_values) - 256)
horizon = 256

context = vib_values[:context_len]
actual_future = vib_values[context_len:context_len + horizon]

print(f"  Context: {context_len} points, Horizon: {horizon} points")
print(f"  Running forecast...")

point_forecast, quantile_forecast = model.forecast(
    horizon=horizon,
    inputs=[context],
)
predicted = point_forecast[0]
quantiles = quantile_forecast[0]

context_times = vib_1min.index[:context_len]
future_times = vib_1min.index[context_len:context_len + horizon]

fig, ax = plt.subplots(figsize=(18, 6))
ax.plot(context_times[-200:], context[-200:], 'b-', alpha=0.7, linewidth=0.8, label='Historical (context)')
ax.plot(future_times[:len(actual_future)], actual_future, 'g-', alpha=0.8, linewidth=0.8, label='Actual')
ax.plot(future_times[:len(predicted)], predicted, 'r--', alpha=0.9, linewidth=1.2, label='TimesFM Predicted')
if quantiles.shape[1] >= 2:
    q_low = quantiles[:, 1]
    q_high = quantiles[:, -2]
    ax.fill_between(future_times[:len(q_low)], q_low, q_high, alpha=0.15, color='red', label='Confidence interval')
ax.set_title('MPB02 Vibration Magnitude: TimesFM Forecast vs Actual', fontsize=14)
ax.set_xlabel('Time')
ax.set_ylabel('Magnitude')
ax.legend(fontsize=10)
ax.xaxis.set_major_formatter(mdates.DateFormatter('%m-%d %H:%M'))
plt.xticks(rotation=30)
plt.tight_layout()
plt.savefig(os.path.join(OUTPUT_DIR, '1_vibration_forecast.png'), dpi=150)
plt.close()
print("  Saved: 1_vibration_forecast.png\n")

# ============================================================
# Part 5: Downtime frequency forecast
# ============================================================
print("=" * 60)
print("Part 5: Downtime frequency forecast with TimesFM...")
print("=" * 60)

downtime_series = hourly_full['count'].values.astype(float)
dt_context_len = min(len(downtime_series) - 24, 120)
dt_horizon = 24

dt_context = downtime_series[:dt_context_len]
dt_actual = downtime_series[dt_context_len:dt_context_len + dt_horizon]

print(f"  Context: {dt_context_len} hours, Horizon: {dt_horizon} hours")
print(f"  Running forecast...")

dt_point, dt_quantile = model.forecast(
    horizon=dt_horizon,
    inputs=[dt_context],
)
dt_predicted = dt_point[0]

ctx_hours = hourly_full['hour'].iloc[:dt_context_len]
fut_hours = hourly_full['hour'].iloc[dt_context_len:dt_context_len + dt_horizon]

fig, ax = plt.subplots(figsize=(16, 5))
ax.bar(ctx_hours[-48:], dt_context[-48:], width=0.035, alpha=0.6, color='steelblue', label='Historical')
ax.bar(fut_hours[:len(dt_actual)], dt_actual, width=0.035, alpha=0.6, color='green', label='Actual')
ax.plot(fut_hours[:len(dt_predicted)], dt_predicted, 'r-o', markersize=4, linewidth=1.5, label='TimesFM Predicted')
ax.set_title('Hourly Unplanned Downtime Count: Forecast vs Actual', fontsize=14)
ax.set_xlabel('Time')
ax.set_ylabel('Downtime count per hour')
ax.legend(fontsize=10)
ax.xaxis.set_major_formatter(mdates.DateFormatter('%m-%d %H:%M'))
plt.xticks(rotation=30)
plt.tight_layout()
plt.savefig(os.path.join(OUTPUT_DIR, '2_downtime_forecast.png'), dpi=150)
plt.close()
print("  Saved: 2_downtime_forecast.png\n")

# ============================================================
# Part 6: Anomaly detection via forecast residuals
# ============================================================
print("=" * 60)
print("Part 6: Anomaly detection - vibration vs downtime...")
print("=" * 60)

window_size = 512
step_size = 256
anomaly_scores = []

print(f"  Sliding window: size={window_size}, step={step_size}")
print(f"  Total points: {len(vib_values)}")

n_windows = 0
for start in range(0, len(vib_values) - window_size - 128, step_size):
    ctx = vib_values[start:start + window_size]
    fut = vib_values[start + window_size:start + window_size + 128]
    if len(fut) < 128:
        break

    pf, _ = model.forecast(horizon=128, inputs=[ctx])
    pred = pf[0]
    residual = np.abs(fut - pred)
    mean_residual = np.mean(residual)

    mid_time = vib_1min.index[start + window_size + 64]
    anomaly_scores.append({'time': mid_time, 'score': mean_residual})
    n_windows += 1
    if n_windows % 5 == 0:
        print(f"    Processed {n_windows} windows...")

anomaly_df = pd.DataFrame(anomaly_scores)
print(f"  Total windows analyzed: {len(anomaly_df)}")

threshold = anomaly_df['score'].mean() + 2 * anomaly_df['score'].std()
anomaly_df['is_anomaly'] = anomaly_df['score'] > threshold
print(f"  Threshold: {threshold:.4f}")
print(f"  Anomalies detected: {anomaly_df['is_anomaly'].sum()}")

unplanned_times = log_df[
    (log_df['LogState'] == 'Unplanned') &
    (log_df['LogTime'] >= vib_df['time'].min()) &
    (log_df['LogTime'] <= vib_df['time'].max())
]['LogTime']

fig, (ax1, ax2, ax3) = plt.subplots(3, 1, figsize=(18, 12), sharex=True)

ax1.plot(vib_1min.index, vib_1min['value'], 'b-', alpha=0.6, linewidth=0.5)
ax1.set_title('MPB02 Vibration Magnitude (1-min avg)', fontsize=13)
ax1.set_ylabel('Magnitude')
for t in unplanned_times:
    ax1.axvline(x=t, color='red', alpha=0.05, linewidth=0.5)

ax2.plot(anomaly_df['time'], anomaly_df['score'], 'b-', linewidth=1)
ax2.axhline(y=threshold, color='r', linestyle='--', linewidth=1, label=f'Threshold ({threshold:.3f})')
anomalies = anomaly_df[anomaly_df['is_anomaly']]
if not anomalies.empty:
    ax2.scatter(anomalies['time'], anomalies['score'], c='red', s=50, zorder=5, label='Anomaly')
ax2.set_title('Forecast Residual Anomaly Score', fontsize=13)
ax2.set_ylabel('Mean |residual|')
ax2.legend()

for t in unplanned_times:
    ax2.axvline(x=t, color='red', alpha=0.05, linewidth=0.5)

downtime_in_range = hourly_full[
    (hourly_full['hour'] >= vib_df['time'].min()) &
    (hourly_full['hour'] <= vib_df['time'].max())
]
ax3.bar(downtime_in_range['hour'], downtime_in_range['count'], width=0.035, color='orangered', alpha=0.7)
ax3.set_title('Hourly Unplanned Downtime Count', fontsize=13)
ax3.set_ylabel('Count')
ax3.set_xlabel('Time')
ax3.xaxis.set_major_formatter(mdates.DateFormatter('%m-%d %H:%M'))
plt.xticks(rotation=30)

plt.tight_layout()
plt.savefig(os.path.join(OUTPUT_DIR, '3_anomaly_vs_downtime.png'), dpi=150)
plt.close()
print("  Saved: 3_anomaly_vs_downtime.png\n")

# ============================================================
# Part 7: Downtime duration forecast
# ============================================================
print("=" * 60)
print("Part 7: Downtime duration forecast...")
print("=" * 60)

dur_series = hourly_full['total_duration'].values.astype(float)
dur_ctx_len = min(len(dur_series) - 24, 120)
dur_horizon = 24

dur_ctx = dur_series[:dur_ctx_len]
dur_actual = dur_series[dur_ctx_len:dur_ctx_len + dur_horizon]

dur_point, dur_quantile = model.forecast(
    horizon=dur_horizon,
    inputs=[dur_ctx],
)
dur_predicted = dur_point[0]

ctx_h = hourly_full['hour'].iloc[:dur_ctx_len]
fut_h = hourly_full['hour'].iloc[dur_ctx_len:dur_ctx_len + dur_horizon]

fig, ax = plt.subplots(figsize=(16, 5))
ax.bar(ctx_h[-48:], dur_ctx[-48:], width=0.035, alpha=0.6, color='steelblue', label='Historical')
ax.bar(fut_h[:len(dur_actual)], dur_actual, width=0.035, alpha=0.6, color='green', label='Actual')
ax.plot(fut_h[:len(dur_predicted)], dur_predicted, 'r-o', markersize=4, linewidth=1.5, label='TimesFM Predicted')
ax.set_title('Hourly Downtime Duration (min): Forecast vs Actual', fontsize=14)
ax.set_xlabel('Time')
ax.set_ylabel('Duration (minutes)')
ax.legend(fontsize=10)
ax.xaxis.set_major_formatter(mdates.DateFormatter('%m-%d %H:%M'))
plt.xticks(rotation=30)
plt.tight_layout()
plt.savefig(os.path.join(OUTPUT_DIR, '4_downtime_duration_forecast.png'), dpi=150)
plt.close()
print("  Saved: 4_downtime_duration_forecast.png\n")

# ============================================================
# Summary
# ============================================================
print("=" * 60)
print("DONE! All outputs saved to:", OUTPUT_DIR)
print("=" * 60)
print("\nOutput files:")
for f in os.listdir(OUTPUT_DIR):
    fpath = os.path.join(OUTPUT_DIR, f)
    size_kb = os.path.getsize(fpath) / 1024
    print(f"  {f}  ({size_kb:.0f} KB)")

mae = np.mean(np.abs(actual_future[:len(predicted)] - predicted))
print(f"\n--- Vibration forecast MAE: {mae:.4f} ---")
print(f"--- Vibration mean value: {np.mean(vib_values):.4f} ---")
print(f"--- MAE / Mean ratio: {mae / np.mean(vib_values) * 100:.1f}% ---")
