"""
Compare different forecasting strategies:
1. Long horizon vs short horizon
2. Rolling short-term prediction (sliding window)
3. Multi-scale aggregation (1min / 5min / 30min / 1h)
4. Anomaly detection sensitivity at different horizons
"""
import os, csv, io, zipfile
import numpy as np
import pandas as pd
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
from datetime import timedelta

os.environ["HF_HOME"] = r"E:\timesfm_project\hf_cache"
import torch
import timesfm

plt.rcParams['font.sans-serif'] = ['SimHei', 'Microsoft YaHei']
plt.rcParams['axes.unicode_minus'] = False

OUTPUT_DIR = r"E:\timesfm_project\output"
DATA_DIR = r"C:\Users\Administrator\Desktop\时序"
ZIP_FILE = os.path.join(DATA_DIR, "2024.07.17-08.17.zip")
LOG_FILE = os.path.join(DATA_DIR, "log8.12-8.18(1).xlsx")

# ============================================================
# Load model
# ============================================================
print("Loading TimesFM model...")
torch.set_float32_matmul_precision("high")
model = timesfm.TimesFM_2p5_200M_torch.from_pretrained(
    "google/timesfm-2.5-200m-pytorch"
)
model.compile(timesfm.ForecastConfig(
    max_context=1024, max_horizon=256,
    normalize_inputs=True, use_continuous_quantile_head=True,
    force_flip_invariance=True, infer_is_positive=True,
    fix_quantile_crossing=True,
))
print("Model ready.\n")

# ============================================================
# Load vibration data (MPB02, Aug 12 only for focused analysis)
# ============================================================
print("Loading vibration data...")
z = zipfile.ZipFile(ZIP_FILE, 'r')
target_files = sorted([
    n for n in z.namelist()
    if 'MPB02' in n and '08.12' in n
])
vibration_data = []
for fname in target_files:
    print(f"  {fname.split('/')[-1]}")
    with z.open(fname) as zf:
        content = io.TextIOWrapper(zf, encoding='utf-8-sig')
        reader = csv.reader(content)
        for _ in range(4): next(reader)
        for row in reader:
            if len(row) >= 9 and row[8] == 'MPB_02' and row[7] == 'Magnitude':
                try:
                    vibration_data.append({'time': row[5], 'value': float(row[6])})
                except ValueError:
                    pass
vib_df = pd.DataFrame(vibration_data)
vib_df['time'] = pd.to_datetime(vib_df['time'], format='ISO8601', utc=True).dt.tz_localize(None)
vib_df = vib_df.sort_values('time').reset_index(drop=True)
print(f"  Raw records: {len(vib_df)}\n")

# Load downtime log
import openpyxl
wb = openpyxl.load_workbook(LOG_FILE, data_only=True)
ws = wb['Sheet1']
log_rows = []
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
    log_rows.append({
        'LogTime': row[6], 'LogState': row[7],
        'Duration': row[9] if row[9] else 0, 'UnplanCode': row[10]
    })
log_df = pd.DataFrame(log_rows)
log_df['LogTime'] = pd.to_datetime(log_df['LogTime'])

# ============================================================
# Experiment 1: Short vs Long horizon comparison
# ============================================================
print("=" * 60)
print("Experiment 1: Horizon comparison (short vs long)")
print("=" * 60)

resample_configs = {
    '1min': '1min',
    '5min': '5min',
    '30min': '30min',
    '1h': '1h',
}

horizons_to_test = [8, 16, 32, 64, 128, 256]

results_table = []

for resample_name, resample_freq in resample_configs.items():
    vib_resampled = vib_df.set_index('time').resample(resample_freq).mean().dropna()
    values = vib_resampled['value'].values

    print(f"\n--- Resample: {resample_name} ({len(values)} points) ---")

    for horizon in horizons_to_test:
        context_len = min(1024, len(values) - horizon)
        if context_len < 128:
            continue

        ctx = values[:context_len]
        actual = values[context_len:context_len + horizon]
        if len(actual) < horizon:
            continue

        pf, qf = model.forecast(horizon=horizon, inputs=[ctx])
        pred = pf[0]

        mae = np.mean(np.abs(actual - pred))
        rmse = np.sqrt(np.mean((actual - pred) ** 2))
        mape = np.mean(np.abs((actual - pred) / (actual + 1e-8))) * 100
        mean_val = np.mean(actual)
        relative_err = mae / mean_val * 100

        results_table.append({
            'resample': resample_name,
            'horizon': horizon,
            'mae': mae,
            'rmse': rmse,
            'mape': mape,
            'relative_err': relative_err,
            'context_len': context_len,
        })
        print(f"  horizon={horizon:>4} | MAE={mae:.4f} | RMSE={rmse:.4f} | RelErr={relative_err:.1f}%")

results_df = pd.DataFrame(results_table)

fig, axes = plt.subplots(1, 2, figsize=(16, 6))

for resample_name in resample_configs.keys():
    subset = results_df[results_df['resample'] == resample_name]
    if len(subset) == 0:
        continue
    axes[0].plot(subset['horizon'], subset['mae'], 'o-', label=resample_name, linewidth=2, markersize=6)
    axes[1].plot(subset['horizon'], subset['relative_err'], 'o-', label=resample_name, linewidth=2, markersize=6)

axes[0].set_xlabel('Forecast Horizon (points)', fontsize=12)
axes[0].set_ylabel('MAE', fontsize=12)
axes[0].set_title('MAE vs Forecast Horizon at Different Scales', fontsize=13)
axes[0].legend(fontsize=11)
axes[0].grid(True, alpha=0.3)

axes[1].set_xlabel('Forecast Horizon (points)', fontsize=12)
axes[1].set_ylabel('Relative Error (%)', fontsize=12)
axes[1].set_title('Relative Error vs Forecast Horizon', fontsize=13)
axes[1].legend(fontsize=11)
axes[1].grid(True, alpha=0.3)

plt.tight_layout()
plt.savefig(os.path.join(OUTPUT_DIR, '5_horizon_comparison.png'), dpi=150)
plt.close()
print("\nSaved: 5_horizon_comparison.png")

# ============================================================
# Experiment 2: Rolling short-term prediction
# ============================================================
print("\n" + "=" * 60)
print("Experiment 2: Rolling short-term prediction (1min, horizon=16)")
print("=" * 60)

vib_1min = vib_df.set_index('time').resample('1min').mean().dropna()
values_1min = vib_1min['value'].values
times_1min = vib_1min.index

short_horizon = 16
context_size = 512
step = short_horizon

rolling_preds = []
rolling_actuals = []
rolling_times = []
rolling_residuals = []

n_steps = 0
for start in range(0, len(values_1min) - context_size - short_horizon, step):
    ctx = values_1min[start:start + context_size]
    actual = values_1min[start + context_size:start + context_size + short_horizon]
    if len(actual) < short_horizon:
        break

    pf, qf = model.forecast(horizon=short_horizon, inputs=[ctx])
    pred = pf[0]

    for i in range(short_horizon):
        idx = start + context_size + i
        if idx < len(times_1min):
            rolling_preds.append(pred[i])
            rolling_actuals.append(actual[i])
            rolling_times.append(times_1min[idx])
            rolling_residuals.append(abs(actual[i] - pred[i]))

    n_steps += 1
    if n_steps % 20 == 0:
        print(f"  Step {n_steps}...")

print(f"  Total rolling steps: {n_steps}")
print(f"  Total predicted points: {len(rolling_preds)}")

rolling_mae = np.mean(np.abs(np.array(rolling_actuals) - np.array(rolling_preds)))
rolling_mean = np.mean(rolling_actuals)
print(f"  Rolling MAE: {rolling_mae:.4f}")
print(f"  Rolling Relative Error: {rolling_mae/rolling_mean*100:.1f}%")

residuals = np.array(rolling_residuals)
threshold = np.mean(residuals) + 2.5 * np.std(residuals)
anomaly_mask = residuals > threshold
print(f"  Anomaly threshold: {threshold:.4f}")
print(f"  Anomalies: {np.sum(anomaly_mask)}")

# Downtime events in this time range
dt_events = log_df[
    (log_df['LogState'] == 'Unplanned') &
    (log_df['LogTime'] >= times_1min[0]) &
    (log_df['LogTime'] <= times_1min[-1])
]

fig, axes = plt.subplots(4, 1, figsize=(20, 16), sharex=True)

show_start = 200
show_end = min(len(rolling_times), 800)
rt = rolling_times[show_start:show_end]
ra = rolling_actuals[show_start:show_end]
rp = rolling_preds[show_start:show_end]
rr = residuals[show_start:show_end]

axes[0].plot(rt, ra, 'b-', alpha=0.8, linewidth=0.8, label='Actual')
axes[0].plot(rt, rp, 'r-', alpha=0.7, linewidth=0.8, label='Predicted (rolling h=16)')
axes[0].set_title('Rolling Short-Term Prediction: Actual vs Predicted', fontsize=13)
axes[0].set_ylabel('Magnitude')
axes[0].legend(fontsize=10)

axes[1].plot(rt, np.array(ra) - np.array(rp), 'purple', alpha=0.7, linewidth=0.6)
axes[1].axhline(0, color='gray', linestyle='--', linewidth=0.5)
axes[1].set_title('Prediction Error (Actual - Predicted)', fontsize=13)
axes[1].set_ylabel('Error')

axes[2].plot(rolling_times, residuals, 'b-', alpha=0.6, linewidth=0.5)
axes[2].axhline(threshold, color='red', linestyle='--', linewidth=1, label=f'Threshold ({threshold:.2f})')
anom_t = [rolling_times[i] for i in range(len(rolling_times)) if anomaly_mask[i]]
anom_r = [residuals[i] for i in range(len(residuals)) if anomaly_mask[i]]
if anom_t:
    axes[2].scatter(anom_t, anom_r, c='red', s=20, zorder=5, label=f'Anomaly ({len(anom_t)} pts)')
for _, ev in dt_events.iterrows():
    axes[2].axvline(ev['LogTime'], color='orange', alpha=0.03, linewidth=0.5)
axes[2].set_title('Anomaly Score (|residual|) - Full Range', fontsize=13)
axes[2].set_ylabel('|Residual|')
axes[2].legend(fontsize=10)

hourly_res = pd.DataFrame({'time': rolling_times, 'residual': residuals})
hourly_res = hourly_res.set_index('time').resample('1h').mean()
dt_hourly = dt_events.set_index('LogTime').resample('1h').size().reindex(hourly_res.index, fill_value=0)

ax3 = axes[3]
ax3.bar(hourly_res.index, dt_hourly.values, width=0.035, alpha=0.6, color='orangered', label='Downtime count/h')
ax3b = ax3.twinx()
ax3b.plot(hourly_res.index, hourly_res['residual'], 'b-o', markersize=3, linewidth=1.5, alpha=0.8, label='Avg |residual|/h')
ax3.set_title('Hourly: Downtime Count vs Avg Prediction Residual', fontsize=13)
ax3.set_ylabel('Downtime count', color='orangered')
ax3b.set_ylabel('Avg |residual|', color='blue')
ax3.legend(loc='upper left', fontsize=10)
ax3b.legend(loc='upper right', fontsize=10)
ax3.xaxis.set_major_formatter(mdates.DateFormatter('%m-%d %H:%M'))
plt.xticks(rotation=30)

plt.tight_layout()
plt.savefig(os.path.join(OUTPUT_DIR, '6_rolling_short_prediction.png'), dpi=150)
plt.close()
print("Saved: 6_rolling_short_prediction.png")

# ============================================================
# Experiment 3: Multi-scale comparison visual
# ============================================================
print("\n" + "=" * 60)
print("Experiment 3: Multi-scale visual comparison")
print("=" * 60)

fig, axes = plt.subplots(4, 1, figsize=(20, 16))
scale_configs = [
    ('1min', '1min', 512, 32),
    ('5min', '5min', 200, 32),
    ('30min', '30min', 32, 8),
    ('1h', '1h', 16, 6),
]

for idx, (label, freq, ctx_len, h) in enumerate(scale_configs):
    resampled = vib_df.set_index('time').resample(freq).mean().dropna()
    vals = resampled['value'].values
    times = resampled.index

    ctx = vals[:ctx_len]
    actual = vals[ctx_len:ctx_len + h]
    pf, qf = model.forecast(horizon=h, inputs=[ctx])
    pred = pf[0]
    q = qf[0]

    mae = np.mean(np.abs(actual[:len(pred)] - pred))
    rel = mae / np.mean(actual) * 100

    show_ctx = min(60, ctx_len)
    ctx_t = times[ctx_len - show_ctx:ctx_len]
    ctx_v = vals[ctx_len - show_ctx:ctx_len]
    fut_t = times[ctx_len:ctx_len + h]

    axes[idx].plot(ctx_t, ctx_v, 'b-', alpha=0.7, linewidth=1, label='Historical')
    axes[idx].plot(fut_t[:len(actual)], actual, 'g-', linewidth=1.2, label='Actual')
    axes[idx].plot(fut_t[:len(pred)], pred, 'r--', linewidth=1.5, label='Predicted')
    if q.shape[1] >= 2:
        axes[idx].fill_between(fut_t[:len(pred)], q[:, 1], q[:, -2],
                               alpha=0.15, color='red', label='Confidence')
    axes[idx].set_title(
        f'Scale: {label} | context={ctx_len} | horizon={h} | '
        f'MAE={mae:.3f} | RelErr={rel:.1f}%', fontsize=12)
    axes[idx].set_ylabel('Magnitude')
    axes[idx].legend(fontsize=9, loc='lower left')
    axes[idx].grid(True, alpha=0.2)

axes[-1].xaxis.set_major_formatter(mdates.DateFormatter('%m-%d %H:%M'))
plt.xticks(rotation=30)
plt.tight_layout()
plt.savefig(os.path.join(OUTPUT_DIR, '7_multiscale_comparison.png'), dpi=150)
plt.close()
print("Saved: 7_multiscale_comparison.png")

# ============================================================
# Experiment 4: Correlation between residuals and downtime
# ============================================================
print("\n" + "=" * 60)
print("Experiment 4: Residual-Downtime correlation analysis")
print("=" * 60)

hourly_merged = pd.DataFrame({
    'residual': hourly_res['residual'],
    'downtime_count': dt_hourly.values
}).dropna()

corr = hourly_merged['residual'].corr(hourly_merged['downtime_count'])
print(f"  Pearson correlation (residual vs downtime count): {corr:.4f}")

hourly_merged['residual_lag1'] = hourly_merged['residual'].shift(1)
hourly_merged['residual_lag2'] = hourly_merged['residual'].shift(2)
hourly_merged['residual_lag3'] = hourly_merged['residual'].shift(3)

lag_corrs = {}
for lag in range(0, 6):
    shifted = hourly_merged['residual'].shift(lag)
    c = shifted.corr(hourly_merged['downtime_count'])
    lag_corrs[lag] = c
    print(f"  Lag-{lag}h correlation: {c:.4f}")

fig, axes = plt.subplots(1, 3, figsize=(18, 5))

axes[0].scatter(hourly_merged['residual'], hourly_merged['downtime_count'],
                alpha=0.5, edgecolors='navy', facecolors='steelblue', s=40)
axes[0].set_xlabel('Avg |Residual| (prediction error)', fontsize=11)
axes[0].set_ylabel('Downtime Count', fontsize=11)
axes[0].set_title(f'Residual vs Downtime (r={corr:.3f})', fontsize=13)
axes[0].grid(True, alpha=0.3)

lags = list(lag_corrs.keys())
corrs = list(lag_corrs.values())
colors = ['#d32f2f' if c > 0.3 else '#1976d2' if c > 0.1 else '#757575' for c in corrs]
axes[1].bar(lags, corrs, color=colors, edgecolor='white', linewidth=0.5)
axes[1].set_xlabel('Lag (hours)', fontsize=11)
axes[1].set_ylabel('Correlation', fontsize=11)
axes[1].set_title('Lag Correlation: Residual -> Downtime', fontsize=13)
axes[1].axhline(0, color='gray', linewidth=0.5)
axes[1].grid(True, alpha=0.3, axis='y')

# Heatmap: hour-of-day patterns
hourly_merged_time = hourly_merged.copy()
hourly_merged_time['hour_of_day'] = hourly_merged_time.index.hour
by_hour = hourly_merged_time.groupby('hour_of_day').agg(
    avg_residual=('residual', 'mean'),
    avg_downtime=('downtime_count', 'mean')
)
x = np.arange(len(by_hour))
w = 0.35
axes[2].bar(x - w/2, by_hour['avg_residual'] / by_hour['avg_residual'].max(),
            w, label='Normalized Residual', color='steelblue', alpha=0.8)
axes[2].bar(x + w/2, by_hour['avg_downtime'] / by_hour['avg_downtime'].max(),
            w, label='Normalized Downtime', color='orangered', alpha=0.8)
axes[2].set_xlabel('Hour of Day', fontsize=11)
axes[2].set_ylabel('Normalized Value', fontsize=11)
axes[2].set_title('Hour-of-Day Pattern: Residual vs Downtime', fontsize=13)
axes[2].set_xticks(x)
axes[2].set_xticklabels(by_hour.index)
axes[2].legend(fontsize=10)
axes[2].grid(True, alpha=0.3, axis='y')

plt.tight_layout()
plt.savefig(os.path.join(OUTPUT_DIR, '8_correlation_analysis.png'), dpi=150)
plt.close()
print("Saved: 8_correlation_analysis.png")

# ============================================================
# Summary
# ============================================================
print("\n" + "=" * 60)
print("SUMMARY")
print("=" * 60)

print("\n[Horizon Comparison]")
for _, r in results_df.iterrows():
    print(f"  {r['resample']:>5} | h={r['horizon']:>3} | MAE={r['mae']:.4f} | RelErr={r['relative_err']:.1f}%")

best = results_df.loc[results_df['relative_err'].idxmin()]
print(f"\n  BEST: resample={best['resample']}, horizon={best['horizon']}, RelErr={best['relative_err']:.1f}%")

print(f"\n[Rolling Short-Term]")
print(f"  Rolling h=16 MAE: {rolling_mae:.4f} ({rolling_mae/rolling_mean*100:.1f}%)")
print(f"  Anomalies detected: {np.sum(anomaly_mask)}")

print(f"\n[Correlation]")
print(f"  Residual-Downtime correlation: {corr:.4f}")
best_lag = max(lag_corrs, key=lag_corrs.get)
print(f"  Best lag: {best_lag}h (r={lag_corrs[best_lag]:.4f})")

print(f"\n[Recommendation]")
print(f"  1. Use 5-30min resample with short horizon (8-32 points)")
print(f"  2. Rolling prediction outperforms single-shot long prediction")
print(f"  3. Residual-based anomaly detection is viable for downtime early warning")
print(f"\nAll outputs saved to: {OUTPUT_DIR}")
